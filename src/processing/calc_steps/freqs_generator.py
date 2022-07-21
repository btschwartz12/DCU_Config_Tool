from __future__ import generator_stop
from dataclasses import dataclass, fields
import json
from pprint import pformat
from openpyxl import load_workbook
from config.config import Config


from src.utils.utils import StepData

SHEET_NAME = "CustomerFreqs"

REQUIRED_KEYS = ["Customer ID", "Customer Name", "Frequency", "Frequency Use"]

class UnassignedFrequencyException(Exception):
    def __init__(self, freqs):
        self.freqs = freqs

@dataclass
class FrequencyData(StepData):
    """This contains the frequency values provided by the excel sheet.
    ALl frequencies will be loaded, but there is a possibility that some are
    not included in the export, depending on if the user indicated weather or not the system 
    included STAR/SRFN devices"""
    # Unassign - warn and don't use
    CUSTOMER_NAME: str = None
    CUSTOMER_ID: int = None
    STAR_F2_Downlink_Frequency_r46: float = None
    SRFN_Outbound_Downlink_Frequency_r47: float = None
    STAR_F1_Uplink_Frequency_r48: float = None
    SRFN_Inbound_Uplink_Frequency_0_r49: float = None
    SRFN_Inbound_Uplink_Frequency_1_r50: float = None
    SRFN_Inbound_Uplink_Frequency_2_r51: float = None
    SRFN_Inbound_Uplink_Frequency_3_r52: float = None
    SRFN_Inbound_Uplink_Frequency_4_r53: float = None
    SRFN_Inbound_Uplink_Frequency_5_r54: float = None
    SRFN_Inbound_Uplink_Frequency_6_r55: float = None
    SRFN_Inbound_Uplink_Frequency_7_r56: float = None
    SRFN_Inbound_Uplink_Frequency_8_r57: float = None
    SRFN_Inbound_Uplink_Frequency_9_r58: float = None
    SRFN_Inbound_Uplink_Frequency_10_r59: float = None
    SRFN_Inbound_Uplink_Frequency_11_r60: float = None
    SRFN_Inbound_Uplink_Frequency_12_r61: float = None
    SRFN_Inbound_Uplink_Frequency_13_r62: float = None
    SRFN_Inbound_Uplink_Frequency_14_r63: float = None
    num_inbound_frequencies: int = None
    unassigned_frequencies: list[float] = None

    def getInboundFieldNames(self):
        names = []
        for field_data in fields(self):
            if "SRFN_Inbound_Uplink_Frequency" in field_data.name:
                names.append(field_data.name)
        return names

    def getInboundNameByOffset(self, offset):
        inbound_names = self.getInboundFieldNames()
        val = getattr(self, inbound_names[offset])
        if val is None:
            raise Exception("error 810: "+str(inbound_names[offset])+" value is not initialized")
        else:
            return inbound_names[offset]
        


class _FrequencyGenerator:
    """This will get the customer frequencies from the Excel sheet,
    so that they can be used for DTLS calculations"""
    def __init__(self, config: Config):
        self.config = config
        self.frequency_objs = []
        self.unassigned_freqs = []

    def loadFreqsFromJson(self, JSON_FN):
        data = []
        with open(JSON_FN, 'r') as f:
            data = json.load(f)
            if not isinstance(data, list):
                raise Exception("Incompatible data format found: "+str(type(data))+"\nShould be: list")
        for freq_obj in data:
            if not (set(REQUIRED_KEYS) <= set(list(freq_obj.keys()))):
                raise Exception("error 660: imported frequency data entry is missing required keys\n\nrequired keys: "+pformat(REQUIRED_KEYS, indent=2))
        self.frequency_objs = data

    def loadFreqsFromExcel(self, WB_FN, SHEET_NAME):
        """This will use an Excel parser to grab the customer frequencies
        and the data associated with them and return a list containing 
        each frequency data entry"""
        workbook = load_workbook(WB_FN)
        
        if SHEET_NAME not in workbook.sheetnames:
            raise Exception("error 491: invalid sheet name: "+SHEET_NAME)

        sheet = workbook[SHEET_NAME]

        freq_entries = []

        for row in range(2, 18):
            freq_entry = {}

            if sheet['A'+str(row)].value == ' ' or sheet['A'+str(row)].value is None:
                break

            for col in range(0, 12):
                col_letter = chr(col + 65) # converts col number to letter
                key_cell = col_letter+'1'
                key = sheet[key_cell].value

                val_cell = col_letter+str(row)
                val = sheet[val_cell].value

                freq_entry[key] = val
            
            freq_entries.append(freq_entry)

        self.frequency_objs = freq_entries    

    def constructFrequencyData(self) -> FrequencyData:
        """This is called after the import file has been validated and the json data
        has been fetched. This will take the json data and construct a FrequencyData object,
        checking for erros such as multiple outbound frequencies and unassigned frequencies."""
        inbound_freqs = []
        unassigned_freqs = []
        has_f1 = False
        has_f2 = False
        has_outbound = False

        FREQUENCY_DATA = FrequencyData()

        CUSTOMER_NAME = self.frequency_objs[0][self.config.FREQUENCY_KEYS["name"]]
        CUSTOMER_ID = self.frequency_objs[0][self.config.FREQUENCY_KEYS["id"]]
        
        for frequency_obj in self.frequency_objs:
            use = frequency_obj[self.config.FREQUENCY_KEYS["type"]]
            frequency = float(frequency_obj[self.config.FREQUENCY_KEYS["frequency"]])
            name = frequency_obj[self.config.FREQUENCY_KEYS["name"]]
            cust_id = frequency_obj[self.config.FREQUENCY_KEYS["id"]]

            if name != CUSTOMER_NAME:
                raise Exception("error 328: non-matching customer names found in frequency file\n\n"+CUSTOMER_NAME+", "+name)
            if cust_id != CUSTOMER_ID:
                raise Exception("error 328: non-matching customer id's found in frequency file\n\n"+str(CUSTOMER_ID)+", "+str(cust_id))

            if use == 'F1':
                if FREQUENCY_DATA.STAR_F1_Uplink_Frequency_r48 is not None:
                    raise Exception("error 330: multiple F1 frequencies provided ("+str(FREQUENCY_DATA.STAR_F1_Uplink_Frequency_r48)+", "+str(frequency)+")")
                has_f1 = True

                FREQUENCY_DATA.STAR_F1_Uplink_Frequency_r48 = frequency

            elif use == 'F2':
                if FREQUENCY_DATA.STAR_F2_Downlink_Frequency_r46 is not None:
                    raise Exception("error 331: multiple F2 frequencies provided ("+str(FREQUENCY_DATA.STAR_F2_Downlink_Frequency_r46)+", "+str(frequency)+")")
                has_f2 = True

                FREQUENCY_DATA.STAR_F2_Downlink_Frequency_r46 = frequency

            elif use == 'OUTBOUND':
                if FREQUENCY_DATA.SRFN_Outbound_Downlink_Frequency_r47 is not None:
                    raise Exception("error 332: multiple outbound frequencies provided ("+str(FREQUENCY_DATA.SRFN_Outbound_Downlink_Frequency_r47)+", "+str(frequency)+")")
                has_outbound = True

                FREQUENCY_DATA.SRFN_Outbound_Downlink_Frequency_r47 = frequency

            elif use == 'UNASSIGN':
                unassigned_freqs.append(frequency)

            elif use == 'INBOUND':
                inbound_freqs.append(frequency)

        if not has_f1:
            raise Exception("error 340: F1 frequency not provided")
        if not has_f2:
            raise Exception("error 340: F2 frequency not provided")
        if not has_outbound:
            raise Exception("error 340: outbound frequency not provided")


        inbound_field_names = FREQUENCY_DATA.getInboundFieldNames()

        inbound_freqs.sort()

        for i, frequency in enumerate(inbound_freqs):
            setattr(FREQUENCY_DATA, inbound_field_names[i], frequency)


        FREQUENCY_DATA.CUSTOMER_NAME = CUSTOMER_NAME
        FREQUENCY_DATA.CUSTOMER_ID = int(CUSTOMER_ID)
        FREQUENCY_DATA.num_inbound_frequencies = len(inbound_freqs)
        FREQUENCY_DATA.unassigned_frequencies = unassigned_freqs

        return FREQUENCY_DATA
            
    
            

def getFrequencyData(config: Config, FREQUENCIES_FN) -> FrequencyData:
        """This will extract relevent data from the frequencies in the excel tool,
        and create a data structure that stores the F1, F2, Outbound, and sorted Inbound 
        frequencies.
        """

        freq_generator = _FrequencyGenerator(config)

        if FREQUENCIES_FN.endswith('json'):
            freq_generator.loadFreqsFromJson(FREQUENCIES_FN)

        elif FREQUENCIES_FN.endswith('xlsx'):
            freq_generator.loadFreqsFromExcel(FREQUENCIES_FN, SHEET_NAME)

        else:
            print("error 303")

        if freq_generator.frequency_objs == []:
            raise Exception("error 391: failed to load frequency objects. Please check import files")
                
        FREQUENCY_DATA: FrequencyData = freq_generator.constructFrequencyData()

        return FREQUENCY_DATA